from __future__ import annotations

import json
import math
import re
import shutil
from pathlib import Path
from typing import Any

import openpyxl


SOURCE_ROOT = Path(r"D:\360MoveData\Users\Administrator\Documents\New project 17\homepage_category_collection")
PROJECT_ROOT = Path(__file__).resolve().parents[1]
PUBLIC_ROOT = PROJECT_ROOT / "public"
ASSET_ROOT = PUBLIC_ROOT / "assets" / "homepage-catalog"
OUT_FILE = PROJECT_ROOT / "src" / "data" / "taoProductCatalog.ts"


H = {
    "id": "商品ID",
    "shop": "店铺",
    "title": "商品标题",
    "level1": "一级分类",
    "level2": "二级分类",
    "level3": "三级分类/筛选",
    "type": "商品类型",
    "material": "核心材质",
    "form": "款式/器型",
    "recommended": "是否建议上首页",
    "price": "现价",
    "old_price": "原价",
    "currency": "币种",
    "meaning": "首页卖点标题",
    "description": "消费者描述",
    "tags": "推荐标签",
    "main_image": "主图",
    "detail_images": "细节图",
    "summary": "原始描述摘要",
    "params": "参数",
}


DROP_WORDS = (
    "购物袋",
    "福袋",
    "补差",
    "定金",
    "邮费",
    "运费",
    "链接专拍",
    "直播间",
    "工具",
    "包装",
    "礼盒",
    "礼袋",
    "配件",
    "DIY材料/配件",
)


PARENT_LABELS = {
    "bracelet": ("手串", "Bracelets"),
    "necklace": ("项链", "Necklaces"),
    "incense": ("合香珠", "Incense Beads"),
}


GROUP_META = {
    "bracelet-golden-black": ("金玄手串", "Golden Black Bracelets", "以深色珠体、金属点缀或金银配饰为主，适合希望视觉更沉稳、有力量感的成品手串。"),
    "bracelet-nanhong": ("南红手串", "Nanhong Bracelets", "以南红、火焰红、柿子红等暖红系材质为主，色泽鲜明，适合作为醒目而有礼感的祝福手串。"),
    "bracelet-amber": ("琥珀蜜蜡手串", "Amber Bracelets", "以琥珀、蜜蜡等温润材质为主，呈现柔和金黄、茶色与暖光质感。"),
    "bracelet-jade-crystal": ("玉石水晶手串", "Jade & Crystal Bracelets", "整合和田玉、翡翠、水晶、玛瑙、青金、绿松等玉石矿物系产品，适合清透、雅致的礼物场景。"),
    "bracelet-wood-bodhi": ("木质菩提手串", "Wood & Bodhi Bracelets", "以桃木、檀木、沉香、菩提等木质或植物系材料为主，气质安静、日常。"),
    "bracelet-cinnabar-obsidian": ("朱砂黑曜手串", "Cinnabar & Obsidian Bracelets", "整合朱砂、黑曜石等高辨识度材质，适合偏守护、稳定、深色视觉的手串。"),
    "bracelet-selected": ("精选手串", "Selected Bracelets", "无法完全归入单一材质的精选手串，保留在独立系列中方便浏览。"),
    "necklace": ("项链与吊坠", "Necklaces & Pendants", "项链、吊坠、挂件统一归入项链大类，避免把单件商品拆得过细。"),
}


AROMA_META = {
    "东方香调": ("incense-oriental", "合香珠 · 东方香调", "Oriental Incense Beads", "偏东方木质、药香、寺院感的合香珠系列。"),
    "传统合香": ("incense-classic", "合香珠 · 传统合香", "Classic Incense Beads", "传统合香配伍方向，强调温润、沉静和日常佩戴感。"),
    "宁神香调": ("incense-calm", "合香珠 · 宁神香调", "Calming Incense Beads", "偏静心、安定、睡眠前后场景的合香珠。"),
    "文昌香调": ("incense-focus", "合香珠 · 文昌香调", "Focus Incense Beads", "偏学习、专注、书房场景的合香珠。"),
    "木质香调": ("incense-woody", "合香珠 · 木质香调", "Woody Incense Beads", "以木质、檀香、沉香气息为核心的合香珠。"),
    "清凉香调": ("incense-cool", "合香珠 · 清凉香调", "Cooling Incense Beads", "更轻盈、清透、带清凉感的合香珠。"),
    "花香调": ("incense-floral", "合香珠 · 花香调", "Floral Incense Beads", "带花香、柔和与礼物感的合香珠。"),
    "草本香调": ("incense-herbal", "合香珠 · 草本香调", "Herbal Incense Beads", "偏草本、药草、自然气息的合香珠。"),
    "醒神香调": ("incense-awake", "合香珠 · 醒神香调", "Awakening Incense Beads", "偏提神、清醒、明亮感的合香珠。"),
}


BRACELET_RULES = [
    ("bracelet-golden-black", ("金玄", "玄金", "金银配饰")),
    ("bracelet-nanhong", ("南红", "瓦西", "柿子红", "火焰红", "冰红")),
    ("bracelet-amber", ("琥珀", "蜜蜡", "老蜡", "血珀")),
    (
        "bracelet-jade-crystal",
        (
            "和田玉",
            "白玉",
            "青玉",
            "碧玉",
            "翡翠",
            "玉石",
            "水晶",
            "白水晶",
            "黄水晶",
            "粉水晶",
            "紫水晶",
            "海蓝宝",
            "玛瑙",
            "珍珠",
            "青金石",
            "绿松石",
            "松石",
        ),
    ),
    ("bracelet-wood-bodhi", ("木质", "桃木", "檀木", "沉香", "崖柏", "菩提", "绿檀", "小叶紫檀")),
    ("bracelet-cinnabar-obsidian", ("朱砂", "辰砂", "黑曜石", "曜石")),
]


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def clean(value: Any, limit: int = 180) -> str:
    value_text = re.sub(r"\s+", " ", text(value))
    value_text = re.sub(r"店铺：.*?$|已售：.*?$|保障：.*?$|退货.*?$", "", value_text)
    value_text = value_text.strip(" ，。、；;")
    if len(value_text) > limit:
        value_text = value_text[:limit].rstrip(" ，。、；;") + "..."
    return value_text


def safe_name(value: str) -> str:
    value = re.sub(r'[\\/:*?"<>|]+', "_", value)
    value = re.sub(r"\s+", "_", value).strip("._ ")
    return value[:80] or "product"


def split_paths(value: Any) -> list[str]:
    return [part.strip() for part in re.split(r"[;,，；、\n]", text(value)) if part.strip()]


def copy_asset(rel: str, dest_dir: Path, name: str) -> str | None:
    if not rel:
        return None
    source = SOURCE_ROOT / rel.replace("/", "\\")
    if not source.exists():
        return None
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / f"{name}{source.suffix.lower()}"
    if not dest.exists() or dest.stat().st_size != source.stat().st_size:
        shutil.copy2(source, dest)
    return "/" + dest.relative_to(PUBLIC_ROOT).as_posix()


def short_name(row: dict[str, Any]) -> str:
    title = text(row.get(H["title"]))
    material = text(row.get(H["material"]))
    form = text(row.get(H["form"]))
    quoted = re.search(r"[【「](.*?)[】」]", title)
    if quoted and 2 <= len(quoted.group(1).strip()) <= 18:
        return quoted.group(1).strip()

    candidate = re.sub(r"^(青城山珠宝旗舰店|青城山菩福|菩福旗舰店|青城山道香馆)", "", title)
    candidate = re.sub(r"(天然|正品|官方|保真|新款|男女款|女款|男款|送礼|礼物|手串|手链|项链|吊坠|挂坠|非遗|古法)", "", candidate)
    candidate = re.split(r"[ ，,。/｜|]+", candidate)[0].strip(" -_")
    if len(candidate) < 2:
        candidate = f"{material}{form}".strip()
    if len(candidate) < 2:
        candidate = "东方灵饰"
    return candidate[:18]


def should_drop(row: dict[str, Any]) -> bool:
    level1 = text(row.get(H["level1"]))
    level2 = text(row.get(H["level2"]))
    product_type = text(row.get(H["type"]))
    title = text(row.get(H["title"]))
    title_blob = " ".join(text(row.get(key)) for key in (H["title"], H["level1"], H["level2"], H["type"], H["material"], H["form"]))

    if level1 in {"配件工具", "其他配饰"}:
        return True
    if level2 == "DIY材料/配件" or product_type == "DIY材料/配件":
        return True
    if any(word in title_blob for word in ("购物袋", "福袋", "补差", "定金", "邮费", "运费", "链接专拍", "直播间", "包装", "礼盒", "礼袋")):
        return True
    if not title:
        return True
    return False


def classify(row: dict[str, Any]) -> tuple[str, str, str, str, str] | None:
    if should_drop(row):
        return None

    level1 = text(row.get(H["level1"]))
    level2 = text(row.get(H["level2"]))
    product_type = text(row.get(H["type"]))
    material = text(row.get(H["material"]))
    shape_blob = " ".join(
        text(row.get(key))
        for key in (H["title"], H["level1"], H["level2"], H["level3"], H["type"], H["material"], H["form"])
    )

    if level1 == "香饰香珠" or product_type == "合香香饰":
        aroma_key = level2 if level2 in AROMA_META else "东方香调"
        group_id, title_cn, subtitle, description = AROMA_META[aroma_key]
        return ("incense", group_id, title_cn, subtitle, description)

    if level1 == "项链吊坠" or product_type in {"项链", "吊坠/挂件"}:
        title_cn, subtitle, description = GROUP_META["necklace"]
        return ("necklace", "necklace", title_cn, subtitle, description)

    if level1 == "手链手串" or product_type == "手链/手串":
        source_key = " ".join([level2, material, shape_blob])
        for group_id, keywords in BRACELET_RULES:
            if any(keyword in source_key for keyword in keywords):
                title_cn, subtitle, description = GROUP_META[group_id]
                return ("bracelet", group_id, title_cn, subtitle, description)
        title_cn, subtitle, description = GROUP_META["bracelet-selected"]
        return ("bracelet", "bracelet-selected", title_cn, subtitle, description)

    return None


def workbook_path() -> Path:
    candidates = sorted(SOURCE_ROOT.glob("*.xlsx"))
    if not candidates:
        raise FileNotFoundError("No xlsx catalog found")
    return candidates[0]


def load_rows() -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(workbook_path(), read_only=True, data_only=True)
    sheet_name = next(name for name in wb.sheetnames if "商品" in name or "上架" in name)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [text(v) for v in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:]]


def main() -> None:
    rows = load_rows()
    groups: dict[str, dict[str, Any]] = {}
    product_count = 0

    for row in rows:
        if text(row.get(H["recommended"])) != "是":
            continue
        classified = classify(row)
        if classified is None:
            continue

        parent, group_id, group_title, subtitle, group_description = classified
        product_id = text(row.get(H["id"])) or str(product_count + 1)
        display_name = short_name(row)
        dest_dir = ASSET_ROOT / f"{product_id}_{safe_name(display_name)}"

        images: list[str] = []
        main_image = copy_asset(text(row.get(H["main_image"])), dest_dir, "main")
        if main_image:
            images.append(main_image)
        for idx, rel in enumerate(split_paths(row.get(H["detail_images"]))[:8], 1):
            copied = copy_asset(rel, dest_dir, f"detail_{idx:02d}")
            if copied and copied not in images:
                images.append(copied)
        if not images:
            continue

        product_count += 1
        try:
            price = round(float(row.get(H["price"]) or row.get(H["old_price"]) or 0), 2)
        except Exception:
            price = 0

        item = {
            "id": f"product-{product_id}",
            "code": str(product_count).zfill(3),
            "name": display_name,
            "sourceName": text(row.get(H["title"])),
            "type": text(row.get(H["type"])) or text(row.get(H["level1"])),
            "meaning": clean(row.get(H["meaning"]), 90) or group_title,
            "description": clean(row.get(H["description"]), 190) or clean(row.get(H["summary"]), 190) or text(row.get(H["title"])),
            "details": [detail for detail in [clean(row.get(H["params"]), 180), clean(row.get(H["summary"]), 220)] if detail],
            "image": images[0],
            "images": images,
            "tags": [
                tag
                for tag in [
                    text(row.get(H["material"])),
                    text(row.get(H["level2"])),
                    text(row.get(H["type"])),
                ]
                if tag and tag != "未注明"
            ][:5],
            "price": price,
            "currency": text(row.get(H["currency"])) or "CNY",
        }

        groups.setdefault(
            group_id,
            {
                "id": group_id,
                "parent": parent,
                "title": group_title,
                "subtitle": subtitle,
                "description": group_description,
                "items": [],
            },
        )
        groups[group_id]["items"].append(item)

    parent_order = {"bracelet": 0, "necklace": 1, "incense": 2}
    preferred_order = [
        "bracelet-golden-black",
        "bracelet-nanhong",
        "bracelet-amber",
        "bracelet-jade-crystal",
        "bracelet-wood-bodhi",
        "bracelet-cinnabar-obsidian",
        "bracelet-selected",
        "necklace",
        *[meta[0] for meta in AROMA_META.values()],
    ]
    order_map = {group_id: index for index, group_id in enumerate(preferred_order)}
    catalog = sorted(
        groups.values(),
        key=lambda group: (parent_order.get(group["parent"], 99), order_map.get(group["id"], 999), group["title"]),
    )

    source = (
        "export type TaoCatalogItem = {\n"
        "  id: string;\n"
        "  code: string;\n"
        "  name: string;\n"
        "  sourceName: string;\n"
        "  type: string;\n"
        "  meaning: string;\n"
        "  description: string;\n"
        "  details: string[];\n"
        "  image: string;\n"
        "  images: string[];\n"
        "  tags: string[];\n"
        "  price: number;\n"
        "  currency: string;\n"
        "};\n\n"
        "export type TaoCatalogGroup = {\n"
        "  id: string;\n"
        "  parent: \"bracelet\" | \"necklace\" | \"incense\";\n"
        "  title: string;\n"
        "  subtitle: string;\n"
        "  description: string;\n"
        "  items: TaoCatalogItem[];\n"
        "};\n\n"
        f"export const taoProductCatalog: TaoCatalogGroup[] = {json.dumps(catalog, ensure_ascii=False, indent=2)};\n"
    )
    OUT_FILE.write_text(source, encoding="utf-8")
    summary = {group["title"]: len(group["items"]) for group in catalog}
    print(json.dumps({"groups": len(catalog), "products": product_count, "summary": summary}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
